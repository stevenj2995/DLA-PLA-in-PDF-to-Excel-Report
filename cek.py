"""Memeriksa hasil Excel terhadap PDF sumbernya.

Dipakai setelah satu batch selesai, untuk memastikan isi Excel benar-benar
sama dengan isi dokumennya. Berdiri sendiri: tidak memanggil apa pun dari
Backend/, melainkan membaca ulang PDF dengan cara sendiri. Kalau pembaca di
Backend/ salah, pembaca di sini tidak ikut salah, dan selisihnya ketahuan.

    python cek.py "PDF Files" hasil.xlsx

Dua pemeriksaan berjalan. Yang pertama membandingkan tiap sel dengan hasil
pembacaan ulang. Yang kedua tidak bergantung pada pembacaan sama sekali:
dokumen menulis "IDR X x P% = Y", jadi Your Share dibagi Definite Claim
Amount wajib jatuh di persentase bulat. Baris yang mencampur dua advice
akan memberi rasio yang janggal.
"""
from __future__ import annotations
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import fitz
import openpyxl

CURRENCIES = r"IDR|USD|SGD|EUR|AUD|JPY|GBP|Rp"
RE_LABEL = re.compile(r"^\s*([A-Za-z][A-Za-z0-9 ./&'%()\-]{1,38}?)\s*:\s*(.*)$")
RE_JUNK = re.compile(r"(?i)page \d+ of \d+|for and on behalf of")
RE_GLYPH = re.compile(r"^[A-Z]{40,}$")
RE_SIGN_DATE = re.compile(r"^\s*,?\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}\s*$")
RE_CLOSING = re.compile(
    r"^[A-Z][A-Za-z .]{2,24},\s+(?:\d{1,2}\s+\w+\s+\d{4}|\w+\s+\d{1,2},\s+\d{4})\s*$")
RE_AMOUNT = re.compile(r"(?:" + CURRENCIES + r")\s*\(?\s*([\d.,]*\d)", re.I)
RE_CURRENCY = re.compile(r"\b(" + CURRENCIES + r")\b", re.I)
RE_NUMBER = re.compile(r"\d[\d.,]*\d|\d")

TITLES = ("DEFINITE LOSS ADVICE", "PRELIMINARY LOSS ADVICE")
OWNER = "Reinsurer"
OWNER_NAMES = ("astra buana", "asuransi astra buana", "asuransi astra")


def lines_of(page) -> list[str]:
    bands: dict[int, list[tuple[float, str]]] = {}
    for x0, y0, _x1, _y1, word, *_ in page.get_text("words"):
        bands.setdefault(round(y0 / 3.0), []).append((x0, word))
    return [" ".join(w for _, w in sorted(items)) for _, items in sorted(bands.items())]


def pairs_of(lines: list[str]) -> dict[str, str]:
    """Label -> nilai, dengan baris sambungan disatukan dan kaki surat dibuang."""
    found: dict[str, str] = {}
    last: str | None = None
    for raw in lines:
        line = " ".join(raw.split())
        if not line or RE_GLYPH.match(line) or RE_SIGN_DATE.match(line) or RE_JUNK.search(line):
            last = None
            continue
        m = RE_LABEL.match(line)
        if not m:
            if RE_CLOSING.match(line):
                break
            if last:
                found[last] = f"{found[last]} {line}".strip()
            continue
        last = " ".join(m.group(1).split())
        found[last] = m.group(2).strip()
    return found


def astra_advice(path: Path) -> dict[str, str] | None:
    """Advice di berkas ini yang ditujukan ke Astra Buana, kalau ada."""
    with fitz.open(path) as pdf:
        for page in pdf:
            lines = lines_of(page)
            head = " ".join(lines[:6]).upper()
            if not any(t in head for t in TITLES):
                continue
            found = pairs_of(lines)
            who = (found.get(OWNER) or "").casefold()
            if any(n in who for n in OWNER_NAMES):
                return found
    return None


def currency(text: str) -> str:
    m = RE_CURRENCY.search(text or "")
    return m.group(1).upper() if m else ""


def amount(text: str) -> str:
    m = RE_AMOUNT.search(text or "")
    return m.group(1) if m else ""


def last_number(text: str) -> str:
    found = RE_NUMBER.findall(text or "")
    return found[-1] if found else ""


def without_money(text: str) -> str:
    return re.split(r"\b(?:" + CURRENCIES + r")\b", text or "",
                    maxsplit=1, flags=re.I)[0].strip(" :-,")


# kolom Excel -> (label sumber, cara mengambil nilainya)
RULES: dict[str, tuple[str, object]] = {
    "Insured Interest": ("Insured Interest", without_money),
    "Total Sum Insured": ("Total Sum Insured", amount),
    "Definite Claim Amount": ("Nett Amount", amount),
    "Your Share on Nett Loss": ("Your Share on Nett Loss", last_number),
}
# kolom Currency mengambil mata uang dari kolom nominal tepat setelahnya
MONEY_AFTER_CURRENCY = ("Total Sum Insured", "Definite Claim Amount",
                        "Your Share on Nett Loss")


def read_sheet(path: Path) -> list[tuple[list[str], list[list]]]:
    """Tiap tabel di sheet sebagai (header, baris)."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    starts = [r for r in range(1, ws.max_row + 1)
              if str(ws.cell(r, 1).value or "").startswith("Tabel")] or [0]
    starts.append(ws.max_row + 1)
    tables = []
    for i in range(len(starts) - 1):
        head_at = starts[i] + 1 if starts[i] else 1
        headers = [ws.cell(head_at, c).value for c in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h]
        rows = []
        for r in range(head_at + 1, starts[i + 1]):
            values = [ws.cell(r, j + 1).value for j in range(len(headers))]
            if any(v not in (None, "") for v in values):
                rows.append(values)
        if headers and rows:
            tables.append((headers, rows))
    return tables


def compare(folder: Path, workbook: Path) -> int:
    pdfs = sorted(folder.rglob("*.pdf"))
    expected = {p.name: a for p in pdfs if (a := astra_advice(p)) is not None}
    tables = read_sheet(workbook)

    # Header dan nilai disimpan berpasangan, bukan sebagai dict: kolom Currency
    # muncul beberapa kali di satu tabel, dan dict akan membuang kembarannya.
    in_sheet: dict[str, tuple[list[str], list]] = {}
    for headers, rows in tables:
        if "Sumber PDF" not in headers:
            continue
        at = headers.index("Sumber PDF")
        for values in rows:
            name = values[at]
            if name:
                in_sheet[str(name)] = (headers, values)

    print(f"PDF diperiksa        : {len(pdfs)}")
    print(f"punya advice Astra   : {len(expected)}")
    print(f"baris di Excel       : {len(in_sheet)}")
    print(f"tabel di Excel       : {len(tables)}")
    print()

    missing = sorted(set(expected) - set(in_sheet))
    extra = sorted(set(in_sheet) - set(expected))
    print(f"KURANG (ada advice Astra, tak masuk Excel) : {len(missing)}")
    for name in missing[:10]:
        print(f"    {name}")
    print(f"LEBIH  (masuk Excel, tak punya advice Astra): {len(extra)}")
    for name in extra[:10]:
        print(f"    {name}")
    print()

    checked = 0
    diffs: dict[str, list] = defaultdict(list)
    for name, source in expected.items():
        entry = in_sheet.get(name)
        if entry is None:
            continue
        headers, values = entry
        for i, header in enumerate(headers):
            if header == "Sumber PDF":
                continue
            if header == "Currency":
                nxt = headers[i + 1] if i + 1 < len(headers) else ""
                label = RULES.get(nxt, (nxt, None))[0]
                want = currency(source.get(label, "")) if nxt in MONEY_AFTER_CURRENCY else ""
            elif header in RULES:
                label, how = RULES[header]
                want = how(source.get(label, "")) if label in source else ""
            else:
                want = source.get(header, "") if header in source else ""
            got = "" if values[i] is None else str(values[i]).strip()
            checked += 1
            if str(want).strip() != got:
                diffs[header].append((name, want, got))

    print(f"nilai dibandingkan   : {checked}")
    total_diff = sum(len(v) for v in diffs.values())
    print(f"beda                 : {total_diff}")
    for header, items in sorted(diffs.items(), key=lambda x: -len(x[1])):
        print(f"  {header}: {len(items)}")
        for name, want, got in items[:3]:
            print(f"      {name[:52]}")
            print(f"        pdf   : {str(want)[:66]!r}")
            print(f"        excel : {str(got)[:66]!r}")

    print()
    ratios = Counter()
    odd = []
    for name, (headers, values) in in_sheet.items():
        def cell(header):
            return values[headers.index(header)] if header in headers else None
        nett, share = cell("Definite Claim Amount"), cell("Your Share on Nett Loss")
        try:
            a = float(re.sub(r"[^\d.]", "", str(nett).replace(",", "")))
            b = float(re.sub(r"[^\d.]", "", str(share).replace(",", "")))
        except (TypeError, ValueError):
            continue
        if not a:
            continue
        pct = round(b / a * 100, 4)
        ratios[pct] += 1
        if abs(pct - round(pct, 2)) > 1e-9 or not 0 < pct <= 100:
            odd.append((name, pct))
    print(f"uji rasio share      : {sum(ratios.values())} baris")
    print(f"rasio janggal        : {len(odd)}")
    for name, pct in odd[:8]:
        print(f"    {pct}%  {name[:56]}")
    if ratios:
        print("  persentase yang muncul:",
              ", ".join(f"{p}% x{n}" for p, n in ratios.most_common(6)))

    print()
    ok = not missing and not extra and not total_diff and not odd
    print("HASIL: BERSIH" if ok else "HASIL: ADA YANG PERLU DILIHAT")
    return 0 if ok else 1


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    folder, workbook = Path(sys.argv[1]), Path(sys.argv[2])
    if not folder.is_dir():
        print(f"Folder tidak ada: {folder}")
        return 2
    if not workbook.is_file():
        print(f"Excel tidak ada: {workbook}")
        return 2
    return compare(folder, workbook)


if __name__ == "__main__":
    sys.exit(main())
