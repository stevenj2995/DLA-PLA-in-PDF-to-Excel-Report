from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_NAME = "Report"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CAPTION_FONT = Font(bold=True, size=12, color="1F3864")
MIN_WIDTH, MAX_WIDTH = 10, 52
GAP = 2  # blank rows between tables


def write(path: str | Path, tables) -> Path:
    """One sheet holding one table per set of parameters.

    Advices that carry different parameters are kept in separate tables rather
    than merged into one wide one. Merging means the columns become the union of
    everything seen, so a document missing half of them leaves a row of blanks
    and the sheet turns unreadable.

    Every value is written as text on purpose: it is lifted from the PDF exactly
    as printed, so Excel cannot reinterpret a number or a date into something
    the letter never said.
    """
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    widest = 0
    at = 1
    for n, table in enumerate(tables):
        headers, rows = table.headers, table.rows
        widest = max(widest, len(headers))

        if len(tables) > 1:
            cell = ws.cell(row=at, column=1, value=f"Tabel {n + 1} - {table.caption}")
            cell.font = CAPTION_FONT
            at += 1

        head_at = at
        for i, name in enumerate(headers, start=1):
            cell = ws.cell(row=at, column=i, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        at += 1

        for row in rows:
            for i, value in enumerate(row, start=1):
                cell = ws.cell(row=at, column=i, value="" if value is None else str(value))
                cell.alignment = Alignment(vertical="top")
                cell.number_format = "@"
            at += 1

        # only the first table can carry the sheet's filter and frozen header
        if n == 0:
            ws.freeze_panes = ws.cell(row=head_at + 1, column=1).coordinate
            last = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A{head_at}:{last}{head_at + len(rows)}"
        at += GAP

    _fit_columns(ws, tables, widest)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path


def _fit_columns(ws, tables, widest: int) -> None:
    """Width per column position, measured across every table on the sheet."""
    for i in range(1, widest + 1):
        longest = 0
        for table in tables:
            if i <= len(table.headers):
                longest = max(longest, len(str(table.headers[i - 1])))
            for row in table.rows:
                if i <= len(row):
                    longest = max(longest, len(str(row[i - 1])))
        ws.column_dimensions[get_column_letter(i)].width = max(
            MIN_WIDTH, min(MAX_WIDTH, longest + 2))
