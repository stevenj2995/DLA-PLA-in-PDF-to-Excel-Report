from __future__ import annotations
import os
import re
import unicodedata
import zipfile
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from .. import settings

# reading the template structure
@dataclass
class Column:
    index: int # 1-based
    letter: str # "A", "AB", ...
    group: str | None # merged header on row 2
    header: str | None # header on row 3 -- what parameters are matched against
    flag: str | None # "Mandatory" / "Optional"
    role: str = "empty"
    number_format: str = "@"

    @property
    def name(self) -> str:
        return (self.header or self.group or self.letter).strip()

    @property
    def clean_name(self) -> str:
        n = self.name.split("\n")[0]
        return " ".join(n.replace("(YYYY-MM-DD)", "").replace("(HH:MM)", "").replace("(Y/N)", "").split()).strip()

    @property
    def full_name(self) -> str:
        if self.group and self.header and self.group != self.header:
            return f"{self.group} - {self.clean_name}"
        return self.clean_name


@dataclass
class Schema:
    columns: list[Column] = field(default_factory=list)

    def __iter__(self):
        return iter(self.columns)

    def __len__(self):
        return len(self.columns)

    @property
    def match_targets(self) -> list[Column]:
        return [k for k in self.columns if k.role == "from_pdf"]

    def summary(self) -> dict[str, int]:
        out: dict[str, int] = {}
        for k in self.columns:
            out[k.role] = out.get(k.role, 0) + 1
        return out


def _role_of(letter: str) -> str:
    if letter == settings.ROW_NUMBER_COLUMN:
        return "row_number"
    if letter in settings.CONSTANT_COLUMNS:
        return "constant"
    if letter in settings.FORMULA_COLUMNS:
        return "formula"
    if letter in settings.COLUMNS_FROM_PDF:
        return "from_pdf"
    if letter == settings.LETTER_DATE_COLUMN:
        return "letter_date"
    if letter in settings.MONITORING_COLUMNS:
        return "monitoring"
    if letter in settings.FEE_COLUMNS:
        return "fee"
    return "empty"


@lru_cache(maxsize=4)
def load_schema(template_path: str | None = None) -> Schema:
    path = Path(template_path) if template_path else settings.template_file()
    wb = openpyxl.load_workbook(path, read_only=False)
    ws = wb[settings.MAIN_SHEET]

    group: dict[int, str] = {}
    for m in ws.merged_cells.ranges:
        if m.min_row == settings.GROUP_ROW:
            value = ws.cell(settings.GROUP_ROW, m.min_col).value
            for c in range(m.min_col, m.max_col + 1):
                if value is not None:
                    group[c] = str(value).strip()

    columns: list[Column] = []
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        g = group.get(c)
        if g is None:
            v = ws.cell(settings.GROUP_ROW, c).value
            g = str(v).strip() if v is not None else None
        h = ws.cell(settings.HEADER_ROW, c).value
        f = ws.cell(settings.FLAG_ROW, c).value
        fmt = ws.cell(settings.FIRST_DATA_ROW, c).number_format or "@"
        columns.append(Column(
            index=c, letter=letter, group=g,
            header=str(h).strip() if h is not None else None,
            flag=(str(f).strip() or None) if f is not None else None,
            role=_role_of(letter), number_format=fmt,
        ))
    wb.close()
    return Schema(columns)

RE_EXTLST = re.compile(r"<extLst>.*?</extLst>", re.S)
RE_XR_UID = re.compile(r'\s+xr:uid="[^"]*"')
RE_X14_DV = re.compile(r"<x14:dataValidation\b")
RE_X14_F = re.compile(r"<xm:f>(.*?)</xm:f>", re.S)
RE_SQREF = re.compile(r"<xm:sqref>(.*?)</xm:sqref>", re.S)
RE_RANGE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)(?::\$?([A-Z]{1,3})\$?(\d+))?")
RE_LIST_SOURCE = re.compile(
    r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+):\$?[A-Z]+\$?(\d+)$")


@dataclass
class Row:
    values: dict[str, object] = field(default_factory=dict) # keyed by column letter
    source: str = ""  # file name of the PDF it came from
    warnings: list[str] = field(default_factory=list)

def _main_sheet_part(zf: zipfile.ZipFile) -> str:
    wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
    m = re.search(
        r'<sheet[^>]*name="%s"[^>]*r:id="(rId\d+)"' % re.escape(settings.MAIN_SHEET),
        wb_xml,
    )
    if not m:
        return "xl/worksheets/sheet1.xml"
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    m2 = re.search(r'<Relationship[^>]*Id="%s"[^>]*Target="([^"]+)"' % m.group(1), rels)
    if not m2:
        return "xl/worksheets/sheet1.xml"
    target = m2.group(1).lstrip("/")
    return target if target.startswith("xl/") else f"xl/{target}"

def _x14_entries(block: str) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for chunk in block.split("</x14:dataValidation>"):
        source = RE_X14_F.search(chunk)
        sqref = RE_SQREF.search(chunk)
        if source and sqref:
            out.append((source.group(1), sqref.group(1)))
    return out


def _columns_of(sqref: str) -> list[str]:
    out: list[str] = []
    for start, _r1, end, _r2 in RE_RANGE.findall(sqref):
        for letter in (start, end or start):
            if letter and letter not in out:
                out.append(letter)
    return out


def _last_row(block: str) -> int:
    rows: list[int] = []
    for sqref in RE_SQREF.findall(block):
        for _start, r1, _end, r2 in RE_RANGE.findall(sqref):
            rows.append(int(r1))
            if r2:
                rows.append(int(r2))
    return max(rows) if rows else 1048576

def _widen_sqref(block: str) -> str:
    entries = _x14_entries(block)
    if not entries:
        return block
    claimed: set[str] = set()
    for _source, sqref in entries:
        columns = _columns_of(sqref)
        if claimed.intersection(columns):
            return block # two lists share a column
        claimed.update(columns)

    first, last = settings.FIRST_DATA_ROW, _last_row(block)

    def widen(m: "re.Match[str]") -> str:
        columns = _columns_of(m.group(1))
        if not columns:
            return m.group(0)
        return "<xm:sqref>%s</xm:sqref>" % " ".join(
            "%s%d:%s%d" % (c, first, c, last) for c in columns)

    return RE_SQREF.sub(widen, block)


def _read_extlst(template_path: Path, part: str) -> str | None:
    with zipfile.ZipFile(template_path) as z:
        xml = z.read(part).decode("utf-8")
    m = RE_EXTLST.search(xml)
    if not m:
        return None
    return _widen_sqref(RE_XR_UID.sub("", m.group(0)))


_TYPOGRAPHIC = {
    "\u2013": "-", "\u2014": "-", "\u2212": "-",
    "\u2018": "'", "\u2019": "'", "\u201c": '"', "\u201d": '"',
    "\u00a0": " ",
}

def canonical_key(value) -> str:
    s = unicodedata.normalize("NFKC", str(value))
    for odd, plain in _TYPOGRAPHIC.items():
        s = s.replace(odd, plain)
    return " ".join(s.split()).casefold()


@lru_cache(maxsize=4)
def dropdown_values(template_path: str) -> dict[str, dict[str, str]]:
    """column letter -> {flattened value: the master list's own spelling}"""
    path = Path(template_path)
    with zipfile.ZipFile(path) as z:
        block = _read_extlst(path, _main_sheet_part(z))
    if not block:
        return {}

    wb = openpyxl.load_workbook(path)
    try:
        out: dict[str, dict[str, str]] = {}
        for source, sqref in _x14_entries(block):
            m = RE_LIST_SOURCE.match(source.strip())
            if not m or m.group(1) not in wb.sheetnames:
                continue
            sheet = wb[m.group(1)]
            column = column_index_from_string(m.group(2))
            allowed: dict[str, str] = {}
            for r in range(int(m.group(3)), int(m.group(4)) + 1):
                v = sheet.cell(r, column).value
                if v not in (None, ""):
                    allowed.setdefault(canonical_key(v), str(v))
            for letter in _columns_of(sqref):
                out[letter] = allowed
        return out
    finally:
        wb.close()


RE_WORD = re.compile(r"[a-z0-9]+")


def _words(key: str) -> set[str]:
    return set(RE_WORD.findall(key))

def resolve_dropdown(allowed: dict[str, str], value) -> str | None:
    if not isinstance(value, str) or not value.strip():
        return None
    key = canonical_key(value)
    if key in allowed:
        return allowed[key]
    words = _words(key)
    if not words:
        return None
    hits = {v for k, v in allowed.items() if words <= _words(k)}
    return hits.pop() if len(hits) == 1 else None


def _fit_dropdown(allowed: dict[str, str], value, column: Column, row: int, corrected: list[str], invalid: list[str]):
    if not isinstance(value, str) or value.startswith("N/A"):
        return value
    exact = resolve_dropdown(allowed, value)
    if exact is None:
        invalid.append("%s%d (%s): '%s' tidak ada di daftar pilihan" % (column.letter, row, column.clean_name, value))
        return value
    if exact != value:
        corrected.append("%s%d (%s): '%s' -> '%s'" % (column.letter, row, column.clean_name, value, exact))
    return exact


def _patch_extlst(source: Path, destination: Path, part: str, block: str) -> None:
    with zipfile.ZipFile(settings.long_path(source)) as zin, \
         zipfile.ZipFile(settings.long_path(destination), "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == part:
                s = RE_EXTLST.sub("", data.decode("utf-8"))
                s = s.replace("</worksheet>", block + "</worksheet>")
                data = s.encode("utf-8")
            zout.writestr(item, data)

def count_dropdowns(path: Path) -> int:
    with zipfile.ZipFile(settings.long_path(path)) as z:
        xml = z.read(_main_sheet_part(z)).decode("utf-8", "ignore")
    return len(RE_X14_DV.findall(xml))


def write_rows(
    rows: list[Row],
    destination: Path,
    *,
    operator_email: str,
    template_path: Path | None = None,
    schema: Schema | None = None,
    start_number: int = 1,
) -> dict:
    template_path = Path(template_path or settings.template_file())
    schema = schema or load_schema(str(template_path))
    destination = Path(destination)
    os.makedirs(settings.long_path(destination.parent), exist_ok=True)

    with zipfile.ZipFile(template_path) as z:
        block = _read_extlst(template_path, _main_sheet_part(z))
    dropdowns_before = count_dropdowns(template_path)
    lists = dropdown_values(str(template_path))

    wb = openpyxl.load_workbook(template_path)
    ws = wb[settings.MAIN_SHEET]

    fmt = {k.letter: ws.cell(settings.FIRST_DATA_ROW, k.index).number_format for k in schema}

    for r in range(settings.FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    constants = dict(settings.CONSTANT_COLUMNS)
    constants[settings.OPERATOR_EMAIL_COLUMN] = operator_email

    corrected: list[str] = []
    invalid: list[str] = []

    for i, b in enumerate(rows):
        r = settings.FIRST_DATA_ROW + i
        for k in schema:
            cell = ws.cell(r, k.index)
            cell.number_format = fmt.get(k.letter, "@")

            if k.letter == settings.ROW_NUMBER_COLUMN:
                value = str(start_number + i)
            elif k.letter in constants:
                # columns hold the same thing on every row, whatever a document happens to say
                value = constants[k.letter]
            elif k.letter in settings.FORMULA_COLUMNS:
                value = _formula(k.letter, r, b)
            elif k.letter in b.values and k.role != "empty":
                value = b.values[k.letter]
                value = None if value == "" else value
            else:
                value = None

            if (value is not None and k.letter in lists
                    and k.letter not in settings.FREE_TEXT_COLUMNS):
                value = _fit_dropdown(lists[k.letter], value, k, r, corrected, invalid)
            cell.value = value

    mandatory_empty = [
        "%s (%s)" % (k.letter, k.clean_name) for k in schema
        if (k.flag or "").lower() == "mandatory"
        and k.role not in ("empty", "fee", "monitoring")
        and k.letter not in settings.DEFERRED_COLUMNS
        and all(ws.cell(settings.FIRST_DATA_ROW + i, k.index).value in (None, "")
                for i in range(len(rows)))
    ] if rows else []

    try:
        wb.save(settings.long_path(destination))
    except PermissionError as e:
        raise PermissionError(
            "Tidak bisa menulis %s - file itu sedang dibuka di Excel. "
            "Tutup dulu filenya, lalu jalankan ulang." % destination.name) from e
    finally:
        wb.close()

    dropdowns_after = count_dropdowns(destination)
    if block and dropdowns_after < dropdowns_before:
        with zipfile.ZipFile(settings.long_path(destination)) as z:
            part = _main_sheet_part(z)
        temp_path = destination.with_suffix(".tmp.xlsx")
        try:
            _patch_extlst(destination, temp_path, part, block)
            os.replace(settings.long_path(temp_path), settings.long_path(destination))
        finally:
            if os.path.exists(settings.long_path(temp_path)):
                os.remove(settings.long_path(temp_path))
        dropdowns_after = count_dropdowns(destination)

    return {
        "file": str(destination),
        "rows": len(rows),
        "dropdowns_before": dropdowns_before,
        "dropdowns_after": dropdowns_after,
        "dropdowns_intact": dropdowns_after == dropdowns_before,
        "corrected": corrected,
        "invalid": invalid,
        "mandatory_empty": mandatory_empty,
    }


def _formula(letter: str, r: int, b: Row) -> str | None:
    return settings.FORMULA_COLUMNS[letter].format(r=r)
